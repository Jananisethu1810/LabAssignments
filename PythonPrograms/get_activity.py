import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

workbook = openpyxl.load_workbook("ticket.xlsx")
worksheet=workbook['Sheet1']
dv = DataValidation(type="list", formula1='Rules!$T$2:$T$8')
worksheet=worksheet.cell(1, 1, value=None);
dv.add('A1:A100')
worksheet.add_data_validation(dv)
worksheet.cell(1, 1, value=None);
workbook.save("ticket.xlsx");


